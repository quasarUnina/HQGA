import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import intervalCode as code
import numpy as np
import scipy.stats
import statistics as stat


def save_obj(obj, name ):
    f= open('obj/'+ name + '.pkl', 'wb')
    pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name):
    with open('obj/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)

def writeFinalXls(fileName, solution, fitness, evaluation):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Chromosome")
    ws1.cell(column=2, row=1, value="Fitness Value")
    ws1.cell(column=3, row=1, value="Evaluations to obtain best")


    row = 2
    for g in solution:
        ws1.cell(column=1, row=row, value=g)
        row += 1

    row = 2
    for g in fitness:
        ws1.cell(column=2, row=row, value=g)
        row += 1

    row = 2
    for g in evaluation:
        ws1.cell(column=3, row=row, value=g)
        row += 1

    wb.save(filename = fileName)

def writeChromosomeEvolutionXls(fileName, chromosome_evolution):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Chromosome 1")
    ws1.cell(column=2, row=1, value="Chromosome 2")
    ws1.cell(column=3, row=1, value="Chromosome 3")

    row = 2
    for l_sup in chromosome_evolution:
        col=1
        for c in l_sup:
            ws1.cell(column=col, row=row, value=str(c))
            col+=1
        row += 1

    wb.save(filename = fileName)

def writeBestsXls(fileName, bests):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Best Chromosome")
    ws1.cell(column=2, row=1, value="Fitness")
    ws1.cell(column=3, row=1, value="Gray code")

    row = 2
    for l_sup in bests:
        col=1
        for c in l_sup:
            ws1.cell(column=col, row=row, value=str(c))
            col+=1
        row += 1

    wb.save(filename = fileName)

def writeFinalFileXls(fileName, chrs, fits, grays):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Best Chromosome")
    ws1.cell(column=2, row=1, value="Fitness")
    ws1.cell(column=3, row=1, value="Gray code")

    row = 2
    for c in chrs:
        ws1.cell(column=1, row=row, value=str(c))
        row += 1

    row = 2
    for fit in fits:
        ws1.cell(column=2, row=row, value=str(fit))
        row += 1

    row = 2
    for g in grays:
        ws1.cell(column=3, row=row, value=str(g))
        row += 1

    wb.save(filename = fileName)


def readXls(filename, col_name, flagFloat=True):
    wb = load_workbook(filename=filename)
    sheet = wb['Sheet']
    fitness_col=sheet[col_name][1:]
    l=[]
    for cell in fitness_col:
        v=cell.value
        if flagFloat==True:
            v=float(v)
        l.append(v)
    return l

def writeResults(fileName, problems,algs, valuesMatrix):
    wb = Workbook()
    ws1 = wb.active

    i=2
    for a in algs:
        ws1.cell(column=i, row=1, value=a)
        i+=1

    i=2
    for p in problems:
        ws1.cell(column=1, row=i, value=p.name)
        i+=1


    col=2
    for l in valuesMatrix:
        row=2
        for v in l:
            ws1.cell(column=col, row=row, value=v)
            row+=1
        col+=1

    wb.save(filename = fileName)

def plotEvolution(problem, chromosome_evolution):
    fig = plt.figure()

    # Make data.
    X = code.getAllValues(problem.lower_bound, problem.upper_bound, problem.num_bit_code)
    Y = code.getAllValues(problem.lower_bound, problem.upper_bound, problem.num_bit_code)
    Z = []
    for x in X:
        l = []
        for y in Y:
            chr_real = [x, y]
            value = problem.computeFitnessFromReal(chr_real)
            l.append(value)
        Z.append(l)
    X, Y = np.meshgrid(X, Y)
    Z = np.array(Z)
    print(X)
    print(Y)
    print(Z)

    # Plot the surface.
    # surf = ax.plot_surface(X, Y, Z, cmap=cm.coolwarm, linewidth=0, antialiased=False)
    # surf = ax.plot_surface(X, Y, Z, rstride=1, cstride=1,cmap='winter', edgecolor='none')
    # surf = ax.plot_surface(X, Y, Z, rstride=1, cstride=1,cmap='jet', edgecolor='none')
    surf = plt.contour(X, Y, Z)
    fig.colorbar(surf, shrink=0.5, aspect=5)
    i=1
    for l_sup in chromosome_evolution:
        for c in l_sup:
            print(c)
            plt.scatter(c[0], c[1], marker=".")
            plt.annotate(str(i), (c[0], c[1]))
        # Add a color bar which maps values to colors.
        plt.savefig(problem.name+'/plot_' + str(i)+"_"+problem.name + '.png', dpi=300, bbox_inches='tight')
        #plt.show()
        i += 1


def writeOptima(fileName, dict_optima):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Problem name")
    ws1.cell(column=2, row=1, value="Optimum Fitness")

    row = 2
    for c in dict_optima.keys():
        ws1.cell(column=1, row=row, value=str(c))
        ws1.cell(column=2, row=row, value=dict_optima[c])
        row += 1

    wb.save(filename=fileName)

def writeOptimaValues(fileName, dict_optima_gray, dict_optima_real):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Problem name")
    ws1.cell(column=2, row=1, value="Optimum Gray Value")
    ws1.cell(column=3, row=1, value="Optimum Real Value")

    row = 2
    for c in dict_optima_real.keys():
        ws1.cell(column=1, row=row, value=str(c))
        if not dict_optima_gray==None:
            ws1.cell(column=2, row=row, value=str(dict_optima_gray[c]))
        ws1.cell(column=3, row=row, value=str(dict_optima_real[c]))
        row += 1

    wb.save(filename=fileName)


def readOptima(filename):
    dict_optima = {}
    wb = load_workbook(filename=filename)
    sheet = wb['Sheet']

    problem_names = sheet['A'][1:]
    optima = sheet['B'][1:]

    l = 0
    for n in problem_names:
        dict_optima[n.value]=optima[l].value
        l += 1

    return dict_optima

def plotGaussian(name_file,x, x_min, x_max, mean, std):
    x = np.linspace(x_min, x_max, 100)
    y = scipy.stats.norm.pdf(x, mean, std)
    plt.plot(x, y, color='coral')
    plt.grid()
    plt.xlim(x_min, x_max)
    plt.ylim(0, 0.5)
    plt.title('', fontsize=10)

    plt.xlabel('x')
    plt.ylabel('Normal Distribution')

    plt.savefig(name_file+".png")
    plt.show()

def hamming_distance(chr1, chr2):
    count=0
    for i in range(len(chr1)):
        if chr1[i]!=chr2[i]:
            count+=1
    return count

def computeHammingDistance(colgray, problem):
    best_sols=problem.getGrayOptimumValue()
    distances=[]
    if problem.dim ==1:
        for sol in best_sols:
            distances.append(hamming_distance(colgray,  sol))
    if problem.dim==2:
        for sol in best_sols:
            best_opt=''
            for s in sol:
                best_opt+=s
            distances.append(hamming_distance(colgray, best_opt))
    return min(distances)


class device():
    def __init__(self, device, isReal=True,noise_model=None, coupling_map=None, basis_gates=None):
        self.device=device
        self.noise_model=noise_model
        self.coupling_map= coupling_map
        self.basis_gates=basis_gates
        self.real=isReal


#l=[2.097152, 27.262976000000002, 1.1650844444444441, 3.961287111111109, 1.1650844444444441]
#print()
#plotGaussian("prova", l, min(l), max(l), stat.mean(l), stat.stdev(l))